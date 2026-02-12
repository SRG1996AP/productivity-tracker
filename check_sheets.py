import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\edwin\Desktop\Department_Daily_Operational_Tracking.xlsx')
print("Sheet names:", wb.sheetnames)
for sheet in wb.sheetnames:
    print(f"\n=== Sheet: {sheet} ===")
    ws = wb[sheet]
    print(f"Max row: {ws.max_row}, Max column: {ws.max_column}")
    for i, row in enumerate(ws.iter_rows(min_col=1, max_col=10, max_row=25), 1):
        values = [cell.value for cell in row]
        if any(values):
            print(f"Row {i}: {values}")
